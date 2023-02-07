import time
import psycopg2
import itertools
from config_BD import host, user, password, database, port
from get_date import get_date
import openpyxl

attempt = 0

def constat():
    global attempt
    attempt += 1

def get_Deliveries_OZON():
    try:
        date_1 = get_date()
        corrent_date = date_1[0]
        last_day = date_1[1]
        connection = psycopg2.connect(
            database=database,
            user=user,
            password=password,
            host=host,
            port=port
        )
        connection.autocommit = True
        with connection.cursor() as cursor:
            cursor.execute(
                f"""SELECT EXISTS (SELECT * FROM warehouses WHERE ДАТА = '{corrent_date}');"""
            )
            list_curs = cursor.fetchall()
        if list_curs[0][0] == True:
            list_full_stoc_OZON = []
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""SELECT * from warehouses WHERE ДАТА = '{corrent_date}' AND Маркетплейс = 'OZON                          ';"""
                )
                list_remains = cursor.fetchall()
                for iten_remains in list_remains:
                    for item_remains in range(0, len(iten_remains), 19):
                        list_full_stoc_OZON.append(list(itertools.islice(iten_remains, item_remains, item_remains + 19)))
        elif list_curs[0][0] == False:
            list_full_stoc_OZON = []
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""SELECT * from warehouses WHERE ДАТА = '{last_day}' AND Маркетплейс = 'OZON                          ';"""
                )
                list_remains = cursor.fetchall()
                for iten_remains in list_remains:
                    for item_remains in range(0, len(iten_remains), 19):
                        list_full_stoc_OZON.append(list(itertools.islice(iten_remains, item_remains, item_remains + 19)))
        return list_full_stoc_OZON
    except Exception as ex:
        if attempt <= 5:
            time.sleep(30)
            constat()
            get_Deliveries_OZON()
    finally:
        if connection:
            connection.close()

def created_file_OZON_1():
    try:
        date_now = get_date()
        corrent_date = date_now[0]
        data_stock = get_Deliveries_OZON()
        book = openpyxl.Workbook()
        sheet = book.active
        sheet["A1"] = "Макетплейс"
        sheet["B1"] = "Склад маркетплейса"
        sheet["C1"] = "Наименование товара"
        sheet["D1"] = "Текущий остаток на складе маркетплейса"
        sheet["E1"] = "Рассчетные продажи за период"
        sheet["F1"] = "Количество в поставке" 
        sheet["G1"] = "Период на который делается подсорт"
        list_Макетплейс = []
        list_Склад = []
        list_Наименование = []
        list_Текущий_остаток= []
        list_Рассчетные_продажи = []
        list_Количество_в_поставке = []
        list_Период = []
        for item in data_stock:
            if item[7] < 0:
                list_Макетплейс.append(item[0])
                list_Склад.append(item[1])
                list_Наименование.append(item[3])
                list_Текущий_остаток.append(item[4])
                list_Рассчетные_продажи.append(item[6])
                list_Количество_в_поставке.append(int((item[7] * -1) * 1.5))
                list_Период.append("2 недели")
        for index_1_OZON, item_1_OZON in enumerate(list_Макетплейс):
            index_sheet = index_1_OZON + 2
            sheet[f"A{index_sheet}"] = item_1_OZON
            sheet[f"B{index_sheet}"] = list_Склад[index_1_OZON]
            sheet[f"C{index_sheet}"] = list_Наименование[index_1_OZON]
            sheet[f"D{index_sheet}"] = list_Текущий_остаток[index_1_OZON]
            sheet[f"E{index_sheet}"] = list_Рассчетные_продажи[index_1_OZON]
            sheet[f"F{index_sheet}"] = list_Количество_в_поставке[index_1_OZON]
            sheet[f"G{index_sheet}"] = list_Период[index_1_OZON]
        book.save(f"Бланк поставки OZON на 2 недели от {corrent_date}.xlsx")
        book.close()
        return f"Бланк поставки OZON на 2 недели от {corrent_date}.xlsx"
    except Exception as ex:
        if attempt <= 5:
            time.sleep(30)
            constat()
            created_file_OZON_1()

def created_file_OZON_2():
    try:
        date_now = get_date()
        corrent_date = date_now[0]
        data_stock = get_Deliveries_OZON()
        book = openpyxl.Workbook()
        sheet = book.active
        sheet["A1"] = "Макетплейс"
        sheet["B1"] = "Склад маркетплейса"
        sheet["C1"] = "Наименование товара"
        sheet["D1"] = "Текущий остаток на складе маркетплейса"
        sheet["E1"] = "Рассчетные продажи за период"
        sheet["F1"] = "Количество в поставке" 
        sheet["G1"] = "Период на который делается подсорт"
        list_Макетплейс = []
        list_Склад = []
        list_Наименование = []
        list_Текущий_остаток= []
        list_Рассчетные_продажи = []
        list_Количество_в_поставке = []
        list_Период = []
        for item in data_stock:
            if item[9] < 0:
                list_Макетплейс.append(item[0])
                list_Склад.append(item[1])
                list_Наименование.append(item[3])
                list_Текущий_остаток.append(item[4])
                list_Рассчетные_продажи.append(item[8])
                list_Количество_в_поставке.append(int((item[9] * -1) * 1.5))
                list_Период.append("4 недели")
        for index_1_OZON, item_1_OZON in enumerate(list_Макетплейс):
            index_sheet = index_1_OZON + 2
            sheet[f"A{index_sheet}"] = item_1_OZON
            sheet[f"B{index_sheet}"] = list_Склад[index_1_OZON]
            sheet[f"C{index_sheet}"] = list_Наименование[index_1_OZON]
            sheet[f"D{index_sheet}"] = list_Текущий_остаток[index_1_OZON]
            sheet[f"E{index_sheet}"] = list_Рассчетные_продажи[index_1_OZON]
            sheet[f"F{index_sheet}"] = list_Количество_в_поставке[index_1_OZON]
            sheet[f"G{index_sheet}"] = list_Период[index_1_OZON]
        book.save(f"Бланк поставки OZON на 4 недели от {corrent_date}.xlsx")
        book.close()
        return f"Бланк поставки OZON на 4 недели от {corrent_date}.xlsx"
    except Exception as ex:
        if attempt <= 5:
            time.sleep(30)
            constat()
            created_file_OZON_2()

def created_file_OZON_3():
    try:
        date_now = get_date()
        corrent_date = date_now[0]
        data_stock = get_Deliveries_OZON()
        book = openpyxl.Workbook()
        sheet = book.active
        sheet["A1"] = "Макетплейс"
        sheet["B1"] = "Склад маркетплейса"
        sheet["C1"] = "Наименование товара"
        sheet["D1"] = "Текущий остаток на складе маркетплейса"
        sheet["E1"] = "Рассчетные продажи за период"
        sheet["F1"] = "Количество в поставке" 
        sheet["G1"] = "Период на который делается подсорт"
        list_Макетплейс = []
        list_Склад = []
        list_Наименование = []
        list_Текущий_остаток= []
        list_Рассчетные_продажи = []
        list_Количество_в_поставке = []
        list_Период = []
        for item in data_stock:
            if item[11] < 0:
                list_Макетплейс.append(item[0])
                list_Склад.append(item[1])
                list_Наименование.append(item[3])
                list_Текущий_остаток.append(item[4])
                list_Рассчетные_продажи.append(item[10] // 1.3)
                list_Количество_в_поставке.append(int(((item[11] // 1.3) * -1) * 1.5))
                list_Период.append("6 недель")
        for index_1_OZON, item_1_OZON in enumerate(list_Макетплейс):
            index_sheet = index_1_OZON + 2
            sheet[f"A{index_sheet}"] = item_1_OZON
            sheet[f"B{index_sheet}"] = list_Склад[index_1_OZON]
            sheet[f"C{index_sheet}"] = list_Наименование[index_1_OZON]
            sheet[f"D{index_sheet}"] = list_Текущий_остаток[index_1_OZON]
            sheet[f"E{index_sheet}"] = list_Рассчетные_продажи[index_1_OZON]
            sheet[f"F{index_sheet}"] = list_Количество_в_поставке[index_1_OZON]
            sheet[f"G{index_sheet}"] = list_Период[index_1_OZON]
        book.save(f"Бланк поставки OZON на 6 недель от {corrent_date}.xlsx")
        book.close()
        return f"Бланк поставки OZON на 6 недель от {corrent_date}.xlsx"
    except Exception as ex:
        if attempt <= 5:
            time.sleep(30)
            constat()
            created_file_OZON_3()

def created_file_OZON_4():
    try:
        date_now = get_date()
        corrent_date = date_now[0]
        data_stock = get_Deliveries_OZON()
        book = openpyxl.Workbook()
        sheet = book.active
        sheet["A1"] = "Макетплейс"
        sheet["B1"] = "Склад маркетплейса"
        sheet["C1"] = "Наименование товара"
        sheet["D1"] = "Текущий остаток на складе маркетплейса"
        sheet["E1"] = "Рассчетные продажи за период"
        sheet["F1"] = "Количество в поставке" 
        sheet["G1"] = "Период на который делается подсорт"
        list_Макетплейс = []
        list_Склад = []
        list_Наименование = []
        list_Текущий_остаток= []
        list_Рассчетные_продажи = []
        list_Количество_в_поставке = []
        list_Период = []
        for item in data_stock:
            if item[11] < 0:
                list_Макетплейс.append(item[0])
                list_Склад.append(item[1])
                list_Наименование.append(item[3])
                list_Текущий_остаток.append(item[4])
                list_Рассчетные_продажи.append(item[10])
                list_Количество_в_поставке.append(int((item[11] * -1) * 1.5))
                list_Период.append("2 месяца")
        for index_1_OZON, item_1_OZON in enumerate(list_Макетплейс):
            index_sheet = index_1_OZON + 2
            sheet[f"A{index_sheet}"] = item_1_OZON
            sheet[f"B{index_sheet}"] = list_Склад[index_1_OZON]
            sheet[f"C{index_sheet}"] = list_Наименование[index_1_OZON]
            sheet[f"D{index_sheet}"] = list_Текущий_остаток[index_1_OZON]
            sheet[f"E{index_sheet}"] = list_Рассчетные_продажи[index_1_OZON]
            sheet[f"F{index_sheet}"] = list_Количество_в_поставке[index_1_OZON]
            sheet[f"G{index_sheet}"] = list_Период[index_1_OZON]
        book.save(f"Бланк поставки OZON на 2 месяца от {corrent_date}.xlsx")
        book.close()
        return f"Бланк поставки OZON на 2 месяца от {corrent_date}.xlsx"
    except Exception as ex:
        if attempt <= 5:
            time.sleep(30)
            constat()
            created_file_OZON_4()

if __name__ == '__main__':
    created_file_OZON_1()