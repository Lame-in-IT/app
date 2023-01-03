import time
import psycopg2
import itertools
from config_BD import host, user, password, database, port
from get_date import get_date
import openpyxl
import csv

attempt = 0

def constat():
    global attempt
    attempt += 1

def get_Deliveries_WB():
    try:
        date_1 = get_date()
        corrent_date = date_1[0]
        last_day = date_1[1]
        connection = psycopg2.connect(
            database=database,
            user=user,
            password=password,
            host=host,
            port=port
        )
        connection.autocommit = True
        with connection.cursor() as cursor:
            cursor.execute(
                f"""SELECT EXISTS (SELECT * FROM warehouses WHERE ДАТА = '{corrent_date}');"""
            )
            list_curs = cursor.fetchall()
        if list_curs[0][0] == True:
            list_full_stoc_WB = []
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""SELECT * from warehouses WHERE ДАТА = '{corrent_date}' AND Маркетплейс = 'Wildberries';"""
                )
                list_remains = cursor.fetchall()
                for iten_remains in list_remains:
                    for item_remains in range(0, len(iten_remains), 19):
                        list_full_stoc_WB.append(list(itertools.islice(iten_remains, item_remains, item_remains + 19)))
            print(list_full_stoc_WB)
        elif list_curs[0][0] == False:
            list_full_stoc_WB = []
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""SELECT * from warehouses WHERE ДАТА = '{last_day}' AND Маркетплейс = 'Wildberries';"""
                )
                list_remains = cursor.fetchall()
                for iten_remains in list_remains:
                    for item_remains in range(0, len(iten_remains), 19):
                        list_full_stoc_WB.append(list(itertools.islice(iten_remains, item_remains, item_remains + 19)))
        return list_full_stoc_WB
    except Exception as ex:
        print(ex)
        if attempt <= 5:
            time.sleep(30)
            print("Перезапуск (get_Deliveries_WB)")
            constat()
            get_Deliveries_WB()
    finally:
        if connection:
            connection.close()

def created_file_WB_1():
    try:
        date_now = get_date()
        corrent_date = date_now[0]
        data_stock = get_Deliveries_WB()
        book = openpyxl.Workbook()
        sheet = book.active
        sheet["A1"] = "Макетплейс"
        sheet["B1"] = "Склад маркетплейса"
        sheet["C1"] = "Наименование товара"
        sheet["D1"] = "Текущий остаток на складе маркетплейса"
        sheet["E1"] = "Рассчетные продажи за период"
        sheet["F1"] = "Количество в поставке" 
        sheet["G1"] = "Период на который делается подсорт"
        list_Макетплейс = []
        list_Склад = []
        list_Наименование = []
        list_Текущий_остаток= []
        list_Рассчетные_продажи = []
        list_Количество_в_поставке = []
        list_Период = []
        for item in data_stock:
            if item[7] <= 0:
                if item[15] > 0:
                    list_Макетплейс.append(item[0])
                    list_Склад.append(item[1])
                    list_Наименование.append(item[3])
                    list_Текущий_остаток.append(item[4])
                    list_Рассчетные_продажи.append(item[15])
                    list_Количество_в_поставке.append(item[15] * 2)
                    list_Период.append("2 недели")
        for index_1_WB, item_1_WB in enumerate(list_Макетплейс):
            index_sheet = index_1_WB + 2
            sheet[f"A{index_sheet}"] = item_1_WB
            sheet[f"B{index_sheet}"] = list_Склад[index_1_WB]
            sheet[f"C{index_sheet}"] = list_Наименование[index_1_WB]
            sheet[f"D{index_sheet}"] = list_Текущий_остаток[index_1_WB]
            sheet[f"E{index_sheet}"] = list_Рассчетные_продажи[index_1_WB]
            sheet[f"F{index_sheet}"] = list_Количество_в_поставке[index_1_WB]
            sheet[f"G{index_sheet}"] = list_Период[index_1_WB]
        book.save(f"Бланк поставки WB на 1 неделю от {corrent_date}.xlsx")
        book.close()
        return f"Бланк поставки WB на 1 неделю от {corrent_date}.xlsx"
    except Exception as ex:
        print(ex)
        if attempt <= 5:
            time.sleep(30)
            print("Перезапуск (created_file_WB_1)")
            constat()
            created_file_WB_1()

def created_file_WB_2():
    try:
        date_now = get_date()
        corrent_date = date_now[0]
        data_stock = get_Deliveries_WB()
        book = openpyxl.Workbook()
        sheet = book.active
        sheet["A1"] = "Макетплейс"
        sheet["B1"] = "Склад маркетплейса"
        sheet["C1"] = "Наименование товара"
        sheet["D1"] = "Текущий остаток на складе маркетплейса"
        sheet["E1"] = "Рассчетные продажи за период"
        sheet["F1"] = "Количество в поставке" 
        sheet["G1"] = "Период на который делается подсорт"
        list_Макетплейс = []
        list_Склад = []
        list_Наименование = []
        list_Текущий_остаток= []
        list_Рассчетные_продажи = []
        list_Количество_в_поставке = []
        list_Период = []
        for item in data_stock:
            if item[7] <= 0:
                if item[15] > 0:
                    list_Макетплейс.append(item[0])
                    list_Склад.append(item[1])
                    list_Наименование.append(item[3])
                    list_Текущий_остаток.append(item[4])
                    list_Рассчетные_продажи.append(item[6])
                    list_Количество_в_поставке.append(item[6] * 2)
                    list_Период.append("4 недели")
        for index_1_WB, item_1_WB in enumerate(list_Макетплейс):
            index_sheet = index_1_WB + 2
            sheet[f"A{index_sheet}"] = item_1_WB
            sheet[f"B{index_sheet}"] = list_Склад[index_1_WB]
            sheet[f"C{index_sheet}"] = list_Наименование[index_1_WB]
            sheet[f"D{index_sheet}"] = list_Текущий_остаток[index_1_WB]
            sheet[f"E{index_sheet}"] = list_Рассчетные_продажи[index_1_WB]
            sheet[f"F{index_sheet}"] = list_Количество_в_поставке[index_1_WB]
            sheet[f"G{index_sheet}"] = list_Период[index_1_WB]
        book.save(f"Бланк поставки WB на 2 недели от {corrent_date}.xlsx")
        book.close()
        return f"Бланк поставки WB на 2 недели от {corrent_date}.xlsx"
    except Exception as ex:
        print(ex)
        if attempt <= 5:
            time.sleep(30)
            print("Перезапуск (created_file_WB_2)")
            constat()
            created_file_WB_2()

def created_file_WB_3():
    try:
        date_now = get_date()
        corrent_date = date_now[0]
        data_stock = get_Deliveries_WB()
        book = openpyxl.Workbook()
        sheet = book.active
        sheet["A1"] = "Макетплейс"
        sheet["B1"] = "Склад маркетплейса"
        sheet["C1"] = "Наименование товара"
        sheet["D1"] = "Текущий остаток на складе маркетплейса"
        sheet["E1"] = "Рассчетные продажи за период"
        sheet["F1"] = "Количество в поставке" 
        sheet["G1"] = "Период на который делается подсорт"
        list_Макетплейс = []
        list_Склад = []
        list_Наименование = []
        list_Текущий_остаток= []
        list_Рассчетные_продажи = []
        list_Количество_в_поставке = []
        list_Период = []
        for item in data_stock:
            if item[7] <= 0:
                if item[15] > 0:
                    list_Макетплейс.append(item[0])
                    list_Склад.append(item[1])
                    list_Наименование.append(item[3])
                    list_Текущий_остаток.append(item[4])
                    list_Рассчетные_продажи.append(item[17])
                    list_Количество_в_поставке.append(item[17] * 2)
                    list_Период.append("6 недель")
        for index_1_WB, item_1_WB in enumerate(list_Макетплейс):
            index_sheet = index_1_WB + 2
            sheet[f"A{index_sheet}"] = item_1_WB
            sheet[f"B{index_sheet}"] = list_Склад[index_1_WB]
            sheet[f"C{index_sheet}"] = list_Наименование[index_1_WB]
            sheet[f"D{index_sheet}"] = list_Текущий_остаток[index_1_WB]
            sheet[f"E{index_sheet}"] = list_Рассчетные_продажи[index_1_WB]
            sheet[f"F{index_sheet}"] = list_Количество_в_поставке[index_1_WB]
            sheet[f"G{index_sheet}"] = list_Период[index_1_WB]
        book.save(f"Бланк поставки WB на 3 недели от {corrent_date}.xlsx")
        book.close()
        return f"Бланк поставки WB на 3 недели от {corrent_date}.xlsx"
    except Exception as ex:
        print(ex)
        if attempt <= 5:
            time.sleep(30)
            print("Перезапуск (created_file_WB_3)")
            constat()
            created_file_WB_3()

def created_file_WB_4():
    try:
        date_now = get_date()
        corrent_date = date_now[0]
        data_stock = get_Deliveries_WB()
        book = openpyxl.Workbook()
        sheet = book.active
        sheet["A1"] = "Макетплейс"
        sheet["B1"] = "Склад маркетплейса"
        sheet["C1"] = "Наименование товара"
        sheet["D1"] = "Текущий остаток на складе маркетплейса"
        sheet["E1"] = "Рассчетные продажи за период"
        sheet["F1"] = "Количество в поставке" 
        sheet["G1"] = "Период на который делается подсорт"
        list_Макетплейс = []
        list_Склад = []
        list_Наименование = []
        list_Текущий_остаток= []
        list_Рассчетные_продажи = []
        list_Количество_в_поставке = []
        list_Период = []
        for item in data_stock:
            if item[7] <= 0:
                if item[15] > 0:
                    list_Макетплейс.append(item[0])
                    list_Склад.append(item[1])
                    list_Наименование.append(item[3])
                    list_Текущий_остаток.append(item[4])
                    list_Рассчетные_продажи.append(item[6] * 2)
                    list_Количество_в_поставке.append(item[6] * 4)
                    list_Период.append("2 месяца")
        for index_1_WB, item_1_WB in enumerate(list_Макетплейс):
            index_sheet = index_1_WB + 2
            sheet[f"A{index_sheet}"] = item_1_WB
            sheet[f"B{index_sheet}"] = list_Склад[index_1_WB]
            sheet[f"C{index_sheet}"] = list_Наименование[index_1_WB]
            sheet[f"D{index_sheet}"] = list_Текущий_остаток[index_1_WB]
            sheet[f"E{index_sheet}"] = list_Рассчетные_продажи[index_1_WB]
            sheet[f"F{index_sheet}"] = list_Количество_в_поставке[index_1_WB]
            sheet[f"G{index_sheet}"] = list_Период[index_1_WB]
        book.save(f"Бланк поставки WB на 1 месяц от {corrent_date}.xlsx")
        book.close()
        return f"Бланк поставки WB на 1 месяц от {corrent_date}.xlsx"
    except Exception as ex:
        print(ex)
        if attempt <= 5:
            time.sleep(30)
            print("Перезапуск (created_file_WB_4)")
            constat()
            created_file_WB_4()

# ['OZON', 'Екатеринбург', 'Бутылка', 'Бутылка для воды-красная', 9, '2088', 1', 8', 3, 6, 6, 3, 9, 0, '2022-09-25', 0, 9, 2, 8]

if __name__ == '__main__':
    created_file_WB_1()